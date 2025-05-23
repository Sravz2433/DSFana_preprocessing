import streamlit as st
import zipfile
import pandas as pd
from io import BytesIO, StringIO
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from typing import Dict, List, Tuple
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DNAProcessor:
    """Main class for processing DNA structure data"""
    
    def __init__(self):
        self.styles = self._init_styles()
    
    def _init_styles(self) -> Dict:
        """Initialize Excel styles"""
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        
        return {
            'fills': {
                'grey': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
                'blue': PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"),
                'pink': PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
                'green': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            },
            'fonts': {
                'bold': Font(bold=True),
                'blue': Font(color="0813f8", bold=False)
            },
            'alignment': {
                'center': Alignment(horizontal="center", vertical="center")
            },
            'borders': {
                'thin': Border(left=thin, right=thin, top=thin, bottom=thin),
                'thick': Border(left=thick, right=thick, top=thick, bottom=thick)
            }
        }
    
    @st.cache_data
    def parse_fasta(_self, sequence_text: str) -> Tuple[List[str], List[str]]:
        """Parse FASTA file content"""
        sequence_lines = sequence_text.strip().splitlines()
        sequence_ids = []
        sequences = []
        
        for i in range(0, len(sequence_lines), 2):
            if i < len(sequence_lines):
                seq_id = sequence_lines[i].lstrip('>')
                seq = sequence_lines[i+1] if i+1 < len(sequence_lines) else ''
                sequence_ids.append(seq_id)
                sequences.append(seq)
        
        return sequence_ids, sequences
    
    def process_zip_data(self, uploaded_zip) -> Tuple[Dict[str, pd.DataFrame], List[int]]:
        """Process ZIP file and return dataframes"""
        dataframes = {}
        row_counts = []
        
        with zipfile.ZipFile(uploaded_zip, 'r') as zip_ref:
            txt_files = [f for f in zip_ref.namelist() if f.endswith('.txt')]
            
            # Use progress bar for file processing
            progress_bar = st.progress(0)
            
            for idx, file_name in enumerate(txt_files):
                try:
                    with zip_ref.open(file_name) as file:
                        # More efficient data processing
                        data = file.read().decode('utf-8')
                        # Replace spaces with commas in one operation
                        modified_data = data.replace(' ', ',')
                        
                        # Read CSV data
                        df = pd.read_csv(StringIO(modified_data), header=None)
                        df = df.dropna(axis=1, how='all')
                        
                        # Calculate mean and round in one operation
                        base_name = os.path.splitext(os.path.basename(file_name))[0]
                        df[f"avg({base_name})"] = df.mean(axis=1).round(5)
                        df = df.round(5)
                        
                        row_counts.append(len(df))
                        dataframes[base_name] = df
                        
                        # Update progress
                        progress_bar.progress((idx + 1) / len(txt_files))
                        
                except Exception as e:
                    logger.error(f"Error processing {file_name}: {str(e)}")
                    st.error(f"Error processing {file_name}: {str(e)}")
                    continue
            
            progress_bar.empty()
        
        return dataframes, row_counts
    
    def create_worksheet_headers(self, ws, dataframes: Dict[str, pd.DataFrame]):
        """Create headers for worksheet"""
        # Parameters header
        parameters_cell = ws.cell(row=1, column=1, value="PARAMETERS")
        parameters_cell.fill = self.styles['fills']['grey']
        parameters_cell.alignment = self.styles['alignment']['center']
        parameters_cell.font = self.styles['fonts']['bold']
        parameters_cell.border = self.styles['borders']['thick']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        
        # Dataframe headers
        start_col = 3
        for df_name, df in dataframes.items():
            col_count = df.shape[1]
            col_end = start_col + col_count - 1
            
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_end)
            header_cell = ws.cell(row=1, column=start_col)
            header_cell.value = df_name
            header_cell.fill = self.styles['fills']['grey']
            header_cell.alignment = self.styles['alignment']['center']
            header_cell.border = self.styles['borders']['thin']
            
            start_col += col_count
    
    def populate_sequence_data(self, ws, sequence_ids: List[str], sequences: List[str], max_rows: int):
        """Populate sequence data in worksheet"""
        # Column headers
        seq_id_cell = ws.cell(row=2, column=1, value="Sequence ID")
        seq_id_cell.font = self.styles['fonts']['bold']
        seq_id_cell.alignment = self.styles['alignment']['center']
        seq_id_cell.border = self.styles['borders']['thin']
        
        seq_cell = ws.cell(row=2, column=2, value="Sequence")
        seq_cell.font = self.styles['fonts']['bold']
        seq_cell.alignment = self.styles['alignment']['center']
        seq_cell.border = self.styles['borders']['thin']
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        
        # Populate data
        for i in range(max_rows):
            seq_id = sequence_ids[i] if i < len(sequence_ids) else ''
            seq_value = sequences[i] if i < len(sequences) else ''
            
            id_cell = ws.cell(row=i + 3, column=1, value=seq_id)
            id_cell.fill = self.styles['fills']['pink']
            id_cell.border = self.styles['borders']['thin']
            
            val_cell = ws.cell(row=i + 3, column=2, value=seq_value)
            val_cell.fill = self.styles['fills']['blue']
            val_cell.border = self.styles['borders']['thin']
    
    def populate_dataframe_data(self, ws, dataframes: Dict[str, pd.DataFrame]):
        """Populate dataframe data in worksheet"""
        col = 3
        
        # Column headers for dataframes
        for df in dataframes.values():
            n_cols = df.shape[1]
            
            # Number columns (0, 1, 2, ...)
            for i in range(n_cols - 1):
                header_cell = ws.cell(row=2, column=col, value=i)
                header_cell.fill = self.styles['fills']['green']
                header_cell.border = self.styles['borders']['thin']
                col += 1
            
            # Mean column
            mean_cell = ws.cell(row=2, column=col, value="Mean")
            mean_cell.fill = self.styles['fills']['green']
            mean_cell.font = self.styles['fonts']['bold']
            mean_cell.alignment = self.styles['alignment']['center']
            mean_cell.border = self.styles['borders']['thin']
            col += 1
        
        # Populate data values
        start_col = 3
        for df in dataframes.values():
            for i, row in df.iterrows():
                for j, val in enumerate(row):
                    cell = ws.cell(row=i + 3, column=start_col + j, value=val)
                    if j == len(row) - 1:  # Mean column
                        cell.font = self.styles['fonts']['blue']
                    cell.border = self.styles['borders']['thin']
            start_col += df.shape[1]
    
    def create_averages_sheet(self, wb, dataframes: Dict[str, pd.DataFrame], 
                            sequence_ids: List[str], sequences: List[str]):
        """Create the averages worksheet"""
        ws2 = wb.create_sheet("Averages")
        
        # Headers
        headers = ["Sequence ID", "Sequence"] + list(dataframes.keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['fonts']['bold']
            cell.alignment = self.styles['alignment']['center']
            cell.border = self.styles['borders']['thin']
        
        # Data
        max_rows = max([df.shape[0] for df in dataframes.values()])
        for i in range(max_rows):
            # Sequence ID
            cell_id = ws2.cell(row=i+2, column=1, 
                              value=sequence_ids[i] if i < len(sequence_ids) else '')
            cell_id.fill = self.styles['fills']['pink']
            cell_id.border = self.styles['borders']['thin']
            
            # Sequence
            cell_seq = ws2.cell(row=i+2, column=2, 
                               value=sequences[i] if i < len(sequences) else '')
            cell_seq.fill = self.styles['fills']['blue']
            cell_seq.border = self.styles['borders']['thin']
            
            # Mean values
            for col_idx, (name, df) in enumerate(dataframes.items(), start=3):
                if i < df.shape[0]:
                    cell = ws2.cell(row=i+2, column=col_idx, value=df.iloc[i, -1])
                    cell.border = self.styles['borders']['thin']
        
        # Adjust column widths
        for col in ws2.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    def create_excel_report(self, dataframes: Dict[str, pd.DataFrame], 
                          sequence_ids: List[str], sequences: List[str]) -> BytesIO:
        """Create complete Excel report"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Combined Data"
        
        max_rows = max([df.shape[0] for df in dataframes.values()])
        
        # Create first worksheet
        self.create_worksheet_headers(ws1, dataframes)
        self.populate_sequence_data(ws1, sequence_ids, sequences, max_rows)
        self.populate_dataframe_data(ws1, dataframes)
        
        # Create second worksheet
        self.create_averages_sheet(wb, dataframes, sequence_ids, sequences)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

def main():
    """Main Streamlit application"""
    st.set_page_config(
        page_title="DSFAna Parameter Data Processor",
        page_icon="🧬",
        layout="wide"
    )
    
    st.title("🧬 DSFAna Parameter Data Processor")
    st.markdown("Process ZIP files of DNA structure data and generate formatted Excel reports")
    
    # Initialize processor
    processor = DNAProcessor()
    
    # File upload section
    col1, col2 = st.columns(2)
    
    with col1:
        uploaded_zip = st.file_uploader("Upload ZIP file with data files", type="zip")
    
    with col2:
        uploaded_seq = st.file_uploader("Upload sequence.fasta file", type="fasta")
    
    if uploaded_zip and uploaded_seq:
        if st.button("🚀 Process Files", type="primary"):
            try:
                with st.spinner("Processing files..."):
                    # Parse sequence data
                    sequence_text = uploaded_seq.read().decode("utf-8")
                    sequence_ids, sequences = processor.parse_fasta(sequence_text)
                    
                    st.info(f"📊 Found {len(sequence_ids)} sequences")
                    
                    # Process ZIP data
                    dataframes, row_counts = processor.process_zip_data(uploaded_zip)
                    
                    # Validate row counts
                    if len(set(row_counts)) != 1:
                        st.error("❌ Not all files have the same number of rows. Please check your data!")
                        st.stop()
                    
                    st.success(f"✅ Processed {len(dataframes)} data files successfully")
                    
                    # Create Excel report
                    with st.spinner("Generating Excel report..."):
                        excel_output = processor.create_excel_report(dataframes, sequence_ids, sequences)
                    
                    st.success("✅ Processing complete!")
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_output,
                        file_name="dna_structure_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    # Display summary
                    with st.expander("📈 Processing Summary"):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Files Processed", len(dataframes))
                        with col2:
                            st.metric("Sequences", len(sequence_ids))
                        with col3:
                            st.metric("Data Points per File", row_counts[0] if row_counts else 0)
            
            except Exception as e:
                logger.error(f"Processing error: {str(e)}")
                st.error(f"❌ An error occurred during processing: {str(e)}")
                st.info("Please check your input files and try again.")

if __name__ == "__main__":
    main()